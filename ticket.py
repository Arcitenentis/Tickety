import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import sqlite3
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from tkcalendar import DateEntry


def update_date_label():
    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    date_label.config(text=current_date)
    root.after(1000, update_date_label)  # Aktualizuj co sekundę

# Initialize the SQLite database and create tables for tickets and done tickets
def init_db():
    conn = sqlite3.connect('tickets.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS tickets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            content TEXT NOT NULL,
            timestamp TEXT NOT NULL,
            username TEXT NOT NULL,
            location TEXT NOT NULL
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS done_tickets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            content TEXT NOT NULL,
            original_timestamp TEXT NOT NULL,
            done_timestamp TEXT NOT NULL,
            done_comment TEXT,
            username TEXT NOT NULL,
            location TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

    # Initialize the deleted tickets database
    conn = sqlite3.connect('dt.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS dt (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            content TEXT NOT NULL,
            original_timestamp TEXT NOT NULL,
            deletion_timestamp TEXT NOT NULL,
            deletion_comment TEXT,
            username TEXT NOT NULL,
            location TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

# Function to fetch tickets from the database
def fetch_tickets():
    conn = sqlite3.connect('tickets.db')
    cursor = conn.cursor()
    cursor.execute('SELECT username, content, timestamp, location FROM tickets')
    tickets = cursor.fetchall()
    conn.close()
    return tickets

# Function to fetch done tickets from the database
def fetch_done_tickets():
    conn = sqlite3.connect('tickets.db')
    cursor = conn.cursor()
    cursor.execute('SELECT content, original_timestamp, done_timestamp, done_comment, username, location FROM done_tickets')
    done_tickets = cursor.fetchall()
    conn.close()
    return done_tickets

# Function to refresh the Treeview with tickets
def refresh_ticket_list():
    for item in ticket_tree.get_children():
        ticket_tree.delete(item)
    
    tickets = fetch_tickets()
    for ticket in tickets:
        ticket_tree.insert("", tk.END, values=(ticket[0], ticket[1], ticket[2], ticket[3]))

# Function to open a new window for adding a ticket
def open_add_ticket_window():
    add_window = tk.Toplevel(root)
    add_window.title("Add New Ticket")
    add_window.geometry("400x350")

    # Create a text entry widget for the new ticket
    ticket_label = tk.Label(add_window, text="Enter your ticket:")
    ticket_label.pack(pady=10)
    new_ticket_entry = tk.Text(add_window, height=10, width=40)
    new_ticket_entry.pack(pady=10)

    # Create a text entry widget for the location
    location_label = tk.Label(add_window, text="Enter location:")
    location_label.pack(pady=5)
    location_entry = tk.Entry(add_window)
    location_entry.pack(pady=5)

    # Function to save the new ticket and refresh the main window
    def submit_new_ticket():
        username = username_entry.get().strip()
        ticket_content = new_ticket_entry.get("1.0", tk.END).strip()
        location = location_entry.get().strip()
        if username and ticket_content and location:
            save_to_db(username, ticket_content, location)
            messagebox.showinfo("Ticket Submitted", "Ticket has been saved!")
            refresh_ticket_list()
            add_window.destroy()
        else:
            messagebox.showwarning("Input Error", "Please enter all fields.")
    
    # Create the submit button in the new ticket window
    submit_button = tk.Button(add_window, text="Submit", command=submit_new_ticket)
    submit_button.pack(pady=10)

# Function to save ticket to SQLite database
def save_to_db(username, content, location):
    conn = sqlite3.connect('tickets.db')
    cursor = conn.cursor()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor.execute('INSERT INTO tickets (username, content, timestamp, location) VALUES (?, ?, ?, ?)', 
                   (username, content, timestamp, location))
    conn.commit()
    conn.close()

# Function to open a window for deleting a ticket with a comment
def open_delete_ticket_window():
    selected_item = ticket_tree.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select a ticket to delete.")
        return

    delete_window = tk.Toplevel(root)
    delete_window.title("Delete Ticket")
    delete_window.geometry("400x200")

    # Create a text entry widget for the deletion comment
    comment_label = tk.Label(delete_window, text="Enter deletion comment:")
    comment_label.pack(pady=10)
    comment_entry = tk.Text(delete_window, height=5, width=40)
    comment_entry.pack(pady=10)

    # Function to delete the ticket and save the comment
    def submit_deletion():
        deletion_comment = comment_entry.get("1.0", tk.END).strip()
        delete_ticket(deletion_comment)
        delete_window.destroy()

    # Create the submit button in the delete ticket window
    submit_button = tk.Button(delete_window, text="Submit", command=submit_deletion)
    submit_button.pack(pady=10)

# Function to delete a ticket and move it to the deleted tickets database
def delete_ticket(deletion_comment):
    selected_item = ticket_tree.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select a ticket to delete.")
        return

    ticket_values = ticket_tree.item(selected_item, 'values')
    ticket_content = ticket_values[1]
    original_timestamp = ticket_values[2]
    deletion_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    username = ticket_values[0]
    location = ticket_values[3]

    # Move the ticket to the deleted tickets database
    conn = sqlite3.connect('dt.db')
    cursor = conn.cursor()
    cursor.execute('INSERT INTO dt (content, original_timestamp, deletion_timestamp, deletion_comment, username) VALUES (?, ?, ?, ?, ?)', 
                   (ticket_content, original_timestamp, deletion_timestamp, deletion_comment, username, location))
    conn.commit()
    conn.close()

    # Delete the ticket from the main database
    conn = sqlite3.connect('tickets.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM tickets WHERE content = ? AND timestamp = ?', (ticket_content, original_timestamp))
    conn.commit()
    conn.close()

    refresh_ticket_list()
    messagebox.showinfo("Ticket Deleted", "Ticket has been deleted and moved to the deleted tickets database.")

# Function to open a window for marking a ticket as done with a comment
def open_done_ticket_window():
    selected_item = ticket_tree.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select a ticket to mark as done.")
        return

    done_window = tk.Toplevel(root)
    done_window.title("Mark Ticket as Done")
    done_window.geometry("400x200")

    # Create a text entry widget for the done comment
    comment_label = tk.Label(done_window, text="Enter done comment:")
    comment_label.pack(pady=10)
    comment_entry = tk.Text(done_window, height=5, width=40)
    comment_entry.pack(pady=10)

    # Function to mark the ticket as done and save the comment
    def submit_done():
        done_comment = comment_entry.get("1.0", tk.END).strip()
        mark_ticket_done(done_comment)
        done_window.destroy()

    # Create the submit button in the done ticket window
    submit_button = tk.Button(done_window, text="Submit", command=submit_done)
    submit_button.pack(pady=10)

# Function to mark a ticket as done and move it to the done tickets database
def mark_ticket_done(done_comment):
    selected_item = ticket_tree.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select a ticket to mark as done.")
        return

    ticket_values = ticket_tree.item(selected_item, 'values')
    if len(ticket_values) < 4:
        messagebox.showerror("Data Error", "Selected ticket does not have enough data.")
        return

    ticket_content = ticket_values[1]
    original_timestamp = ticket_values[2]
    done_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    username = ticket_values[0]
    location = ticket_values[3]

    # Move the ticket to the done tickets database
    conn = sqlite3.connect('tickets.db')
    cursor = conn.cursor()
    cursor.execute('INSERT INTO done_tickets (username, content, original_timestamp, done_timestamp, done_comment, location) VALUES (?, ?, ?, ?, ?, ?)',
                   (username, ticket_content, original_timestamp, done_timestamp, done_comment, location))
    conn.commit()
    conn.close()

    # Delete the ticket from the main database
    conn = sqlite3.connect('tickets.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM tickets WHERE content = ? AND timestamp = ?', (ticket_content, original_timestamp))
    conn.commit()
    conn.close()

    refresh_ticket_list()
    messagebox.showinfo("Ticket Done", "Ticket has been marked as done and moved to the done tickets database.")

# Function to open a window for viewing done tickets
def open_view_done_tickets_window():
    done_window = tk.Toplevel(root)
    done_window.title("View Done Tickets")
    done_window.geometry("1200x600")

    # Create a Treeview widget to display done tickets
    done_ticket_tree = ttk.Treeview(done_window, columns=("Content", "Original Timestamp", "Done Timestamp", "Done Comment", "Username", "Location"), show="headings", height=15)
    done_ticket_tree.heading("Content", text="Zgłoszenie")
    done_ticket_tree.heading("Original Timestamp", text="Data i czas Zgłoszenia")
    done_ticket_tree.heading("Done Timestamp", text="Data i czas Zamknięcia")
    done_ticket_tree.heading("Done Comment", text="Komentarz")
    done_ticket_tree.heading("Username", text="Wykonawca")
    done_ticket_tree.heading("Location", text="Lokalizacja")
    done_ticket_tree.column("Content", width=300)
    done_ticket_tree.column("Original Timestamp", width=150)
    done_ticket_tree.column("Done Timestamp", width=150)
    done_ticket_tree.column("Done Comment", width=300)
    done_ticket_tree.column("Username", width=150)
    done_ticket_tree.column("Location", width=150)
    done_ticket_tree.pack(pady=20)

    # Fetch and display done tickets
    done_tickets = fetch_done_tickets()
    for ticket in done_tickets:
        done_ticket_tree.insert("", tk.END, values=ticket)

    # Create a Text widget to display the full content of the selected ticket
    text_widget = tk.Text(done_window, wrap=tk.WORD, height=10)
    text_widget.pack(expand=True, fill=tk.BOTH, pady=20)

    # Function to display the full content of the selected ticket
    def display_full_content(_):
        selected_item = done_ticket_tree.selection()
        if selected_item:
            ticket_values = done_ticket_tree.item(selected_item, 'values')
            full_content = f"Ticket Content: {ticket_values[0]}\nOriginal Date & Time: {ticket_values[1]}\nDone Date & Time: {ticket_values[2]}\nDone Comment: {ticket_values[3]}\nUsername: {ticket_values[4]}\nLocation: {ticket_values[5]}"
            text_widget.config(state=tk.NORMAL)
            text_widget.delete(1.0, tk.END)
            text_widget.insert(tk.END, full_content)
            text_widget.config(state=tk.DISABLED)

    # Bind the Treeview selection event to the display_full_content function
    done_ticket_tree.bind("<<TreeviewSelect>>", display_full_content)

# Function to export done tickets to an Excel file
def export_done_tickets_to_excel():
    done_tickets = fetch_done_tickets()
    if not done_tickets:
        messagebox.showwarning("Export Error", "No done tickets to export.")
        return

    # Ask the user to select a directory to save the file
    directory = filedialog.askdirectory()
    if not directory:
        return

    file_path = f"{directory}/Raport Techniczny.xlsx"

    # Group tickets by date
    tickets_by_date = defaultdict(list)
    for ticket in done_tickets:
        date = ticket[2].split(' ')[0]  # Extract date from original timestamp
        tickets_by_date[date].append(ticket)

    # Create a new Excel workbook
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove the default sheet

    # Define styles
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True)

    # Create sheets for each date and add tickets
    for date, tickets in tickets_by_date.items():
        sheet = workbook.create_sheet(title=date)

        # Write headers
        headers = ["Zgłoszenie", "Data i czas Zgłoszenia", "Data i czas Zakończenia", "Komentarz", "Wykonawca"]
        sheet.append(headers)

        # Apply styles to headers
        for cell in sheet[1]:
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border

        # Write tickets
        for ticket in tickets:
            sheet.append(ticket)

        # Apply border and wrap text to all cells
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = header_border
                cell.alignment = wrap_alignment

        # Adjust column widths to a fixed value to control text wrapping
        fixed_width = 40  # Set a fixed width for columns to control wrapping
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = fixed_width

        # Adjust row heights based on text length
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.value:
                    num_lines = len(str(cell.value)) // fixed_width + 1
                    sheet.row_dimensions[cell.row].height = num_lines * 60   # Adjust the multiplier as needed

    # Save the workbook to the specified file
    workbook.save(file_path)

    messagebox.showinfo("Export Successful", f"Done tickets have been exported to {file_path}")

def open_search_window():
    search_window = tk.Toplevel(root)
    search_window.title("Szukaj Zgłoszenia")
    search_window.geometry("400x300")

    # Create a text entry widget for the search query
    search_label = tk.Label(search_window, text="Wprowadź kryteria wyszukiwania:")
    search_label.pack(pady=10)
    search_entry = tk.Entry(search_window, width=40)
    search_entry.pack(pady=10)

    # Create a dropdown menu for selecting the search criteria
    criteria_label = tk.Label(search_window, text="Wybierz kryteria wyszukiwania:")
    criteria_label.pack(pady=10)
    criteria_var = tk.StringVar(search_window)
    criteria_var.set("Lokalizacja")  # Domyślne kryterium
    criteria_menu = tk.OptionMenu(search_window, criteria_var, "Lokalizacja", "Data", "Wykonawca")
    criteria_menu.pack(pady=10)

    # Create date entry widgets for date range selection
    start_date_label = tk.Label(search_window, text="Wybierz datę początkową:")
    start_date_entry = DateEntry(search_window, width=12, background='darkblue', foreground='white', borderwidth=2)
    end_date_label = tk.Label(search_window, text="Wybierz datę końcową:")
    end_date_entry = DateEntry(search_window, width=12, background='darkblue', foreground='white', borderwidth=2)

    # Function to show/hide date entry widgets based on selected criteria
    def update_date_widgets(*args):
        if criteria_var.get() == "Data":
            start_date_label.pack(pady=5)
            start_date_entry.pack(pady=5)
            end_date_label.pack(pady=5)
            end_date_entry.pack(pady=5)
            search_entry.pack_forget()
        else:
            start_date_label.pack_forget()
            start_date_entry.pack_forget()
            end_date_label.pack_forget()
            end_date_entry.pack_forget()
            search_entry.pack(pady=10)

    criteria_var.trace("w", update_date_widgets)

    # Function to perform the search and display results
    def submit_search():
        query = search_entry.get().strip()
        criteria = criteria_var.get()
        if criteria == "Data":
            start_date = start_date_entry.get_date().strftime('%Y-%m-%d')
            end_date = end_date_entry.get_date().strftime('%Y-%m-%d')
            search_results = search_tickets_by_date(start_date, end_date)
        elif query:
            search_results = search_tickets(query, criteria)
        else:
            messagebox.showwarning("Input Error", "Wprowadź kryteria wyszukiwania.")
            return
        display_search_results(search_results)
        search_window.destroy()

    # Create the submit button in the search window
    submit_button = tk.Button(search_window, text="Szukaj", command=submit_search)
    submit_button.pack(pady=10)

def search_tickets(query, criteria):
    conn = sqlite3.connect('tickets.db')
    cursor = conn.cursor()
    
    if criteria == "Lokalizacja":
        cursor.execute('SELECT username, content, done_timestamp AS timestamp, location FROM done_tickets WHERE location LIKE ?', ('%' + query + '%',))
    elif criteria == "Data":
        cursor.execute('SELECT username, content, done_timestamp AS timestamp, location FROM done_tickets WHERE done_timestamp LIKE ?', ('%' + query + '%',))
    elif criteria == "Wykonawca":
        cursor.execute('SELECT username, content, done_timestamp AS timestamp, location FROM done_tickets WHERE username LIKE ?', ('%' + query + '%',))
    
    results = cursor.fetchall()
    conn.close()
    return results

def search_tickets_by_date(start_date, end_date):
    conn = sqlite3.connect('tickets.db')
    cursor = conn.cursor()
    cursor.execute('SELECT username, content, done_timestamp AS timestamp, location FROM done_tickets WHERE done_timestamp BETWEEN ? AND ?', 
                   (start_date, end_date))
    results = cursor.fetchall()
    conn.close()
    return results

def display_search_results(results):
    results_window = tk.Toplevel(root)
    results_window.title("Wyniki Wyszukiwania")
    results_window.geometry("800x400")

    # Create a Treeview widget to display search results
    results_tree = ttk.Treeview(results_window, columns=("Username", "Content", "Timestamp", "Location"), show="headings", height=15)
    results_tree.heading("Username", text="Imie")
    results_tree.heading("Content", text="Zgłoszenie")
    results_tree.heading("Timestamp", text="Data i czas")
    results_tree.heading("Location", text="Lokalizacja")
    results_tree.column("Username", width=150)
    results_tree.column("Content", width=300)
    results_tree.column("Timestamp", width=150)
    results_tree.column("Location", width=150)
    results_tree.pack(pady=20)

    # Insert search results into the Treeview
    for result in results:
        results_tree.insert("", tk.END, values=result)

# Create the main window
root = tk.Tk()
root.title("Ticket Viewer")
root.geometry("900x520")

# Create a label to display the current date
date_label = tk.Label(root, text="", font=("Helvetica", 12))
date_label.pack(pady=5)

# Start updating the date label
update_date_label()

# Create a text entry widget for the username
username_label = tk.Label(root, text="Podaj Imie:")
username_label.pack(pady=5)
username_entry = tk.Entry(root)
username_entry.pack(pady=5)

# Create a Treeview widget to display tickets
ticket_tree = ttk.Treeview(root, columns=("Username", "Content", "Timestamp", "Location"), show="headings", height=15)
ticket_tree.heading("Username", text="Imie")
ticket_tree.heading("Content", text="Zgłoszenie")
ticket_tree.heading("Timestamp", text="Data i czas")
ticket_tree.heading("Location", text="Lokalizacja")
ticket_tree.column("Username", width=150)
ticket_tree.column("Content", width=400)
ticket_tree.column("Timestamp", width=150)
ticket_tree.column("Location", width=150)
ticket_tree.pack(pady=10)

# Create a frame to hold the buttons
button_frame = tk.Frame(root)
button_frame.pack(pady=10, fill=tk.X)

# Add a button to open the "Add Ticket" window
add_ticket_button = tk.Button(button_frame, text="Dodaj Zgłoszenie", command=open_add_ticket_window)
add_ticket_button.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)

# Add a button to delete the selected ticket
delete_ticket_button = tk.Button(button_frame, text="Usuń zgłoszenie", command=open_delete_ticket_window)
delete_ticket_button.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)

# Add a button to mark the selected ticket as done
done_ticket_button = tk.Button(button_frame, text="Zrobione", command=open_done_ticket_window)
done_ticket_button.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)

# Add a button to view done tickets
view_done_tickets_button = tk.Button(button_frame, text="Zobacz Zrobione", command=open_view_done_tickets_window)
view_done_tickets_button.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)

# Add a button to export done tickets to an Excel file
export_done_tickets_button = tk.Button(button_frame, text="Zapisz do Exela", command=export_done_tickets_to_excel)
export_done_tickets_button.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)

# Add a button to search tickets
search_ticket_button = tk.Button(button_frame, text="Szukaj", command=open_search_window)
search_ticket_button.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)

# Initialize the database and fetch tickets
init_db()
refresh_ticket_list()

# Start the Tkinter event loop
root.mainloop()